from common_utils import get_target_order
import pandas as pd
import openpyxl, pprint, os, sys
from datetime import datetime, timedelta
from openpyxl.styles import Alignment


def main():
    # 相対パスを絶対パスに変換する
    base_dir = os.path.dirname(__file__)

    # その上でcsvファイルを読み込む
    input_path = os.path.join(base_dir, '..', 'input', 'products.csv')
    products = pd.read_csv(input_path, encoding="shift_jis")

    input_path = os.path.join(base_dir, '..', 'input', 'orders.csv')
    orders = pd.read_csv(input_path, encoding="shift_jis")

    # # テンプレートを読み込む
    template_path = os.path.join(base_dir, '..', 'template', 'quotation_template.xlsx')
    quotation_wb = openpyxl.load_workbook(template_path)

    # 対象のワークシートを読み込む
    quotation_ws = quotation_wb["Quotation"]

    # IDで指定した注文情報から、消費税、合計金額などを算出しデータにまとめる
    merged, target_id = get_target_order(products, orders)


    # 見積書への入力作業
    # 明細の開始行
    start_row = 19

    # 右詰めスタイルを作成
    right_align = Alignment(horizontal="right")

    # 日付設定
    today = datetime.today()
    due_date = today + timedelta(days=30)

    # 明細行描写
    for i, (_, row) in enumerate(merged.iterrows()):
        # No(連番)
        quotation_ws.cell(row=start_row + i, column=2, value=i + 1)
        # 商品名
        quotation_ws.cell(row=start_row + i, column=3, value=row["product_name"])
         # 数量
        quotation_ws.cell(row=start_row + i, column=4, value=row["quantity"])
        # 単価(円)
        cell = quotation_ws.cell(row=start_row + i, column=5, value=row["unit_price"])
        cell.number_format = '"¥"#,##0'
        # 金額(円)
        cell = quotation_ws.cell(row=start_row + i, column=6, value=row["subtotal"])
        cell.number_format = '"¥"#,##0'

    # 小計金額
    cell = quotation_ws.cell(row=35, column=6, value=merged["subtotal"].sum())
    cell.number_format = '"¥"#,##0'

    # 消費税
    cell = quotation_ws.cell(row=36, column=6, value=merged["VAT"].sum())
    cell.number_format = '"¥"#,##0'

    # 合計金額
    cell = quotation_ws.cell(row=37, column=6, value=merged["total_amount"].sum())
    cell.number_format = '"¥"#,##0'

    # 宛名
    quotation_ws["B3"] = merged.iloc[0]["customer_name"] + "  様"

    # 見積日
    cell = quotation_ws.cell(row=12, column=6, value=today.strftime("%Y年%m月%d日"))
    cell.alignment = right_align

    # 振込期限
    cell = quotation_ws.cell(row=13, column=6, value=due_date.strftime("%Y年%m月%d日"))
    cell.alignment = right_align


    # 作成したファイルを保存
    output_dir = os.path.join(base_dir, '..', 'output')
    os.makedirs(output_dir, exist_ok=True)  # ← フォルダがなければ作成

    filename = f"{target_id}_quotation.xlsx"
    output_path = os.path.join(output_dir, filename)

    quotation_wb.save(output_path)

if __name__ == '__main__':
    main()
