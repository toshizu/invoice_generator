from common_utils import get_target_order
import pandas as pd
import openpyxl, pprint, os, sys
from datetime import datetime, timedelta
from openpyxl.styles import Alignment
from dateutil.relativedelta import relativedelta


def main():
    # 相対パスを絶対パスに変換する
    base_dir = os.path.dirname(__file__)

    # その上でcsvファイルを読み込む
    input_path = os.path.join(base_dir, '..', 'input', 'products.csv')
    products = pd.read_csv(input_path, encoding="shift_jis")

    input_path = os.path.join(base_dir, '..', 'input', 'orders.csv')
    orders = pd.read_csv(input_path, encoding="shift_jis")

    # テンプレートを読み込む
    template_path = os.path.join(base_dir, '..', 'template', 'invoice_template.xlsx')
    invoice_wb = openpyxl.load_workbook(template_path)

    # 対象のワークシートを読み込む
    invoice_ws = invoice_wb["Invoice"]

    # IDで指定した注文情報から、消費税、合計金額などを算出しデータにまとめる
    merged, target_id = get_target_order(products, orders)


    # 請求書への入力作業
    # 明細の開始行
    start_row = 15

    # 右詰めスタイルを作成
    right_align = Alignment(horizontal="right")

    # 日付設定
    today = datetime.today()
    next_month_same_day = today + relativedelta(months=1)

    # 明細行描写
    for i, (_, row) in enumerate(merged.iterrows()):
        # No(連番)
        invoice_ws.cell(row=start_row + i, column=1, value=i + 1)
        # 商品名
        invoice_ws.cell(row=start_row + i, column=2, value=row["product_name"])
         # 数量
        invoice_ws.cell(row=start_row + i, column=3, value=row["quantity"])
        # 単価(円)
        cell = invoice_ws.cell(row=start_row + i, column=4, value=row["unit_price"])
        cell.number_format = '"¥"#,##0'
        # 税率
        invoice_ws.cell(row=start_row + i, column =5, value=str(row["tax_rate"])+"%")
        # 金額(円)
        cell = invoice_ws.cell(row=start_row + i, column=6, value=row["subtotal"])
        cell.number_format = '"¥"#,##0'

    # 小計
    cell = invoice_ws.cell(row=31, column=6, value=merged["subtotal"].sum())
    cell.number_format = '"¥"#,##0'

    # 消費税
    cell = invoice_ws.cell(row=32, column=6, value=merged["VAT"].sum())
    cell.number_format = '"¥"#,##0'

    # 合計金額
    cell = invoice_ws.cell(row=33, column=6, value=merged["total_amount"].sum())
    cell.number_format = '"¥"#,##0'

    # 請求金額
    amount = merged["total_amount"].sum()
    formatted = f"¥{amount:,.0f}(税込)"

    cell = invoice_ws.cell(row=12, column=3, value=formatted)
    cell.alignment = right_align

    # 税率区分
    # 10%
    cell = invoice_ws.cell(row=36, column=4,value=merged[merged["tax_rate"] == 10]["VAT"].sum())
    cell.number_format = '"¥"#,##0'

    cell = invoice_ws.cell(row=36, column=5,value=merged[merged["tax_rate"] == 10]["subtotal"].sum())
    cell.number_format = '"¥"#,##0'
    cell.alignment = right_align

    # 8%
    cell = invoice_ws.cell(row=37, column=4,value=merged[merged["tax_rate"] == 8]["VAT"].sum())
    cell.number_format = '"¥"#,##0'

    cell = invoice_ws.cell(row=37, column=5,value=merged[merged["tax_rate"] == 8]["subtotal"].sum())
    cell.number_format = '"¥"#,##0'
    cell.alignment = right_align

    # 宛名
    invoice_ws["A4"] = merged.iloc[0]["customer_name"] + "  様"

    # 発行日
    cell = invoice_ws.cell(row=5, column=6, value=today.strftime("%Y年%m月%d日"))
    cell.alignment = right_align

    # 登録番号
    cell = invoice_ws.cell(row=6, column=6,value=merged["invoice_registration_number"].iloc[0])
    cell.alignment = right_align

    # 請求番号
    cell = invoice_ws.cell(row=7, column=6,value=merged["order_id"].iloc[0])
    cell.alignment = right_align

    # お支払期日
    cell = invoice_ws.cell(row=42, column=3,value=next_month_same_day.strftime("%Y年%m月末"))
    cell.alignment = right_align


    # 作成したファイルを保存
    output_dir = os.path.join(base_dir, '..', 'output')
    os.makedirs(output_dir, exist_ok=True)  # ← フォルダがなければ作成

    filename = f"{target_id}_invoice.xlsx"
    output_path = os.path.join(output_dir, filename)

    invoice_wb.save(output_path)

if __name__ == '__main__':
    main()
